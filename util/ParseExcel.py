from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font
import locale
import time


class ParseExcel(object):

    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.wb = load_workbook(excel_file_path)
        self.ws = self.wb[self.wb.sheetnames[0]]
        # print(self.ws.title)

    def get_all_sheet_names(self):
        return self.wb.sheetnames

    def get_sheet_name_by_index(self, index):
        return self.wb.sheetnames[index - 1]

    def get_excel_file_path(self):
        return self.excel_file_path

    def create_sheet(self, sheet_name, position=None):

        try:
            if position:
                self.wb.create_sheet(sheet_name, position)
            else:
                self.wb.create_sheet(sheet_name)
            self.save()
            return True
        except Exception as e:
            print(e)
            return False
    def get_sheet_by_name(self,sheet_name):
        return self.wb.get_sheet_by_name(sheet_name)

    def set_sheet_by_name(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            print("%s sheet不存在，请重新设置！" % sheet_name)
            return False
        self.ws = self.wb[sheet_name]
        return True

    def set_sheet_by_index(self, index):
        self.ws = self.wb[self.get_sheet_name_by_index(index)]
        print("设定的sheet名称是：", self.ws.title)

    def get_cell_value(self, row_no, col_no, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        return self.ws.cell(row_no, col_no).value

    def get_row_values(self, row_no, sheet_name=None):
        cell_values = []
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None

        for cell in list(self.ws.rows)[row_no - 1]:
            cell_values.append(cell.value)

        return cell_values

    def get_col_values(self, col_no, sheet_name=None):
        cell_values = []
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        for cell in list(self.ws.columns)[col_no - 1]:
            cell_values.append(cell.value)
        return cell_values

    def get_some_values(self, min_row_no, min_col_no,max_row_no, max_col_no, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        values = []
        for i in range(min_row_no, max_row_no + 1):
            row_values = []
            for j in range(min_col_no, max_col_no + 1):
                row_values.append(self.ws.cell(row=i, column=j).value)
            values.append(row_values)

        return values

    def save(self):
        self.wb.save(self.excel_file_path)

    def write_cell_value(self, row_no, col_no, value, style=None, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return False
        if style is None:
            style = colors.BLACK
        elif style == "red":
            style = colors.RED
        elif style == "green":
            style = colors.DARKGREEN
        self.ws.cell(row=row_no, column=col_no).font = Font(color=style)
        self.ws.cell(row=row_no, column=col_no, value=value)
        self.save()
        return True

    def write_current_time(self, row_no, col_no, style=None, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return False
        if style is None:
            style = colors.BLACK
        elif style == "red":
            style = colors.RED
        elif style == "greed":
            style = colors.GREEN
        locale.setlocale(locale.LC_ALL, 'en')
        locale.setlocale(locale.LC_CTYPE, 'chinese')
        self.ws.cell(row=row_no, column=col_no).font = Font(color=style)
        self.ws.cell(row=row_no, column=col_no,
                     value=time.strftime("%Y年%m月%d日 %H时%M分%S秒"))
        self.save()
        return True


if __name__ == "__main__":
    excel = ParseExcel(r"D:\study\光荣之路\正式课\第十九天\test.xlsx")
    # print(excel.get_excel_file_path())
    # print(excel.get_cell_value(1,1))
    # print(excel.get_cell_value(3,3))
    # excel.set_sheet_by_name("xxxx")
    # excel.set_sheet_by_name("Sheet2")
    # print(excel.get_cell_value(3,3))
    # print(excel.get_cell_value(3,3,"xxx"))
    # print(excel.get_cell_value(3,3,"Sheet2"))
    # print(excel.get_row_values(1))
    # print(excel.get_row_values(1,"Sheet2"))
    # print(excel.get_col_values(1))
    # print(excel.get_col_values(1,"Sheet2"))
    # print(excel.get_some_values(1,1,5,5))
    # print(excel.get_some_values(1,1,3,3,"Sheet2"))
    # print(excel.write_cell_value(6,1,"光荣之路吴老师","red"))
    # print(excel.write_current_time(6,1,"red"))
    # print(excel.get_all_sheet_names())
    # print(excel.get_sheet_name_by_index(1))
    # excel.set_sheet_by_index(2)
    print(excel.create_sheet("光荣之路"))